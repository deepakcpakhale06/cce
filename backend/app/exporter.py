from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from io import BytesIO
from .models import EstimateRow

HEADERS = [
    'Component/Function Name',
    'AWS Service Name',
    'Quantity',
    'Configuration',
    'Assumptions',
    'Cost Per Month (USD)',
    'Yearly Cost (USD)',
]

def create_excel_workbook(rows: list[EstimateRow]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Estimate'

    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = cell.font.copy(bold=True)

    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=row.componentName)
        sheet.cell(row=row_index, column=2, value=row.awsServiceName)
        sheet.cell(row=row_index, column=3, value=row.quantity)
        sheet.cell(row=row_index, column=4, value=row.configuration)
        sheet.cell(row=row_index, column=5, value=row.assumptions)
        sheet.cell(row=row_index, column=6, value=row.costPerMonth)
        sheet.cell(row=row_index, column=7, value=row.yearlyCost)

    for col in range(1, len(HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def streaming_excel_response(rows: list[EstimateRow]) -> StreamingResponse:
    stream = create_excel_workbook(rows)
    return StreamingResponse(
        stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="aws-cost-estimate.xlsx"'},
    )
